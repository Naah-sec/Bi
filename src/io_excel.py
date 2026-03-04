from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


def _table_to_dataframe(excel_path: Path, table_name: str) -> pd.DataFrame | None:
    wb = load_workbook(excel_path, read_only=False, data_only=True)
    try:
        for ws in wb.worksheets:
            for table in ws.tables.values():
                if table.name == table_name:
                    data = ws[table.ref]
                    rows = [[cell.value for cell in row] for row in data]
                    if not rows:
                        return pd.DataFrame()
                    headers = [str(h) if h is not None else "" for h in rows[0]]
                    return pd.DataFrame(rows[1:], columns=headers)
    finally:
        wb.close()
    return None


def _sheet_to_dataframe(excel_path: Path, sheet_name: str) -> pd.DataFrame | None:
    with pd.ExcelFile(excel_path) as xls:
        if sheet_name in xls.sheet_names:
            return pd.read_excel(excel_path, sheet_name=sheet_name)
    return None


def read_table_or_sheet(
    excel_path: Path, table_candidates: list[str], sheet_candidates: list[str], required: bool = True
) -> pd.DataFrame | None:
    for table_name in table_candidates:
        table_df = _table_to_dataframe(excel_path, table_name)
        if table_df is not None:
            return table_df

    for sheet_name in sheet_candidates:
        sheet_df = _sheet_to_dataframe(excel_path, sheet_name)
        if sheet_df is not None:
            return sheet_df

    if required:
        raise FileNotFoundError(
            f"Could not find table/sheet. Tried tables={table_candidates}, sheets={sheet_candidates}"
        )
    return None


def read_sales_excel_inputs(excel_path: Path) -> dict[str, Any]:
    sales_raw = read_table_or_sheet(
        excel_path=excel_path,
        table_candidates=["tblSalesRaw", "SalesRaw"],
        sheet_candidates=["SalesRaw", "tblSalesRaw"],
        required=True,
    )

    map_product = read_table_or_sheet(
        excel_path=excel_path,
        table_candidates=["tblProductCategoryMap", "Map_ProductCategory"],
        sheet_candidates=["Map_ProductCategory", "tblProductCategoryMap"],
        required=False,
    )

    map_typevente = read_table_or_sheet(
        excel_path=excel_path,
        table_candidates=["tblTypeVenteMap", "Map_TypeVente"],
        sheet_candidates=["Map_TypeVente", "tblTypeVenteMap"],
        required=False,
    )

    map_wilaya = read_table_or_sheet(
        excel_path=excel_path,
        table_candidates=["tblWilayaMap", "Map_Wilaya"],
        sheet_candidates=["Map_Wilaya", "tblWilayaMap"],
        required=False,
    )

    return {
        "sales_raw": sales_raw,
        "map_product_category": map_product,
        "map_typevente": map_typevente,
        "map_wilaya": map_wilaya,
    }


def read_workbook_inputs(excel_path: Path) -> dict[str, Any]:
    # Backward-compatible alias used by existing sales code.
    return read_sales_excel_inputs(excel_path)


def read_purchases_excel_inputs(excel_path: Path) -> dict[str, Any]:
    purchases_raw = read_table_or_sheet(
        excel_path=excel_path,
        table_candidates=["tblPurchasesRaw", "PurchasesRaw", "tblSalesRaw", "SalesRaw"],
        sheet_candidates=["PurchasesRaw", "tblPurchasesRaw", "SalesRaw", "tblSalesRaw"],
        required=True,
    )

    map_product = read_table_or_sheet(
        excel_path=excel_path,
        table_candidates=["tblProductCategoryMap", "Map_ProductCategory"],
        sheet_candidates=["Map_ProductCategory", "tblProductCategoryMap"],
        required=False,
    )

    map_typeachat = read_table_or_sheet(
        excel_path=excel_path,
        table_candidates=["tblTypeAchatMap", "Map_TypeAchat"],
        sheet_candidates=["Map_TypeAchat", "tblTypeAchatMap"],
        required=False,
    )

    return {
        "purchases_raw": purchases_raw,
        "map_product_category": map_product,
        "map_typeachat": map_typeachat,
        "source": "excel",
    }


def read_purchases_csv_inputs(csv_path: Path) -> dict[str, Any]:
    if not csv_path.exists():
        raise FileNotFoundError(f"Purchases CSV file not found: {csv_path}")

    purchases_raw = pd.read_csv(csv_path)
    return {
        "purchases_raw": purchases_raw,
        "map_product_category": None,
        "map_typeachat": None,
        "source": "csv",
    }


def read_purchases_inputs(
    purchases_excel_path: Path,
    purchases_csv_path: Path,
    source: str = "auto",
) -> dict[str, Any]:
    normalized = source.strip().lower()
    if normalized not in {"auto", "excel", "csv", "none"}:
        raise ValueError("purchases source must be one of: auto, excel, csv, none")

    if normalized == "none":
        return {
            "purchases_raw": pd.DataFrame(),
            "map_product_category": None,
            "map_typeachat": None,
            "source": "none",
        }

    if normalized == "excel":
        if not purchases_excel_path.exists():
            raise FileNotFoundError(f"Purchases Excel file not found: {purchases_excel_path}")
        return read_purchases_excel_inputs(purchases_excel_path)

    if normalized == "csv":
        return read_purchases_csv_inputs(purchases_csv_path)

    # auto mode
    if purchases_excel_path.exists():
        return read_purchases_excel_inputs(purchases_excel_path)
    if purchases_csv_path.exists():
        return read_purchases_csv_inputs(purchases_csv_path)

    return {
        "purchases_raw": pd.DataFrame(),
        "map_product_category": None,
        "map_typeachat": None,
        "source": "none",
    }
